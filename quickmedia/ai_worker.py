"""Async AI analysis worker for QuickMedia.

Uses a SQLite-backed queue (ai_queue table) to process AI analysis
tasks in the background without blocking file scanning.
"""

import os
from quickmedia.database import Database
from quickmedia.config import Config
from quickmedia.ai import VisionAnalyzer, TextAnalyzer, TranscriptionAnalyzer, OllamaAdapter, EmbeddingAnalyzer, merge_frame_results, extract_video_frames
from quickmedia.prompt_config import PromptConfig
from quickmedia.embedding import _build_field_text, ChromaStore, EmbeddingAdapter, OpenAiEmbeddingAdapter


class AIWorker:
    """Background worker that consumes ai_queue and runs AI analysis."""

    MAX_RETRIES = 3

    def __init__(self, db: Database, config: Config):
        self.db = db
        self.config = config
        timeout = config.get("ai.timeout") or 300
        self._timeout = timeout
        self._prompt_config = PromptConfig(config.config_dir) if hasattr(config, 'config_dir') else None
        self._transcriber = TranscriptionAnalyzer()

        # Load provider registry
        import os as _os
        self._config_dir = config.config_dir if hasattr(config, 'config_dir') else _os.path.expanduser("~/.asset-manager")
        self._env = {}
        env_path = _os.path.join(self._config_dir, ".env")
        if _os.path.isfile(env_path):
            with open(env_path, "r") as f:
                for line in f:
                    line = line.strip()
                    if "=" in line and not line.startswith("#"):
                        k, v = line.split("=", 1)
                        self._env[k] = v
        from quickmedia.providers import ProviderRegistry
        models_path = _os.path.join(self._config_dir, "models.yaml")
        self._registry = ProviderRegistry(config, models_path)


    def _pc(self):
        """Reload PromptConfig from disk on each analysis."""
        from quickmedia.prompt_config import PromptConfig, get_current_language
        return PromptConfig(self._config_dir, get_current_language()) if getattr(self, "_config_dir", None) else None
    def _get_adapter(self, task_type: str):
        """Create an adapter for the given task type based on task_models config."""
        # Re-read config from disk to pick up live changes
        self.config._load()
        from quickmedia.providers import ProviderRegistry
        self._registry = ProviderRegistry(self.config, os.path.join(self._config_dir, "models.yaml"))
        # Also refresh .env
        env_path = os.path.join(self._config_dir, ".env")
        if os.path.isfile(env_path):
            self._env = {}
            with open(env_path, "r") as f:
                for line in f:
                    line = line.strip()
                    if "=" in line and not line.startswith("#"):
                        k, v = line.split("=", 1)
                        self._env[k] = v
        binding = self._registry.get_task_binding(task_type)
        if not binding or not binding.get("provider") or not binding.get("model"):
            if task_type == "aggregation":
                raise RuntimeError("aggregation_no_model")
            # video_vision falls back to vision model
            if task_type == "video_vision":
                binding = self._registry.get_task_binding("vision")
        if not binding:
            # Fall back to ollama provider config
            ollama_url = self.config.get("providers.ollama.url") or "http://localhost:11434"
            ollama_model = self.config.get("task_models.vision.model") or "qwen3.5:9b"
            return OllamaAdapter(ollama_url, ollama_model, self._timeout)
        provider_name = binding["provider"]
        model = binding["model"]
        provider = self._registry.get_provider(provider_name) or {}
        url = provider.get("url", "")
        api_key = self._env.get(f"{provider_name.upper()}_API_KEY", "")
        if provider_name == "ollama":
            api_key = api_key or "ollama"
            return OllamaAdapter(url.rstrip("/"), model, self._timeout)
        else:
            from quickmedia.openai_adapter import OpenAIAdapter
            return OpenAIAdapter(base_url=url, api_key=api_key, model=model, timeout=self._timeout, provider_name=provider_name)

    def enqueue(self, asset_id: int, task_type: str) -> None:
        """Add an AI analysis task to the queue (idempotent)."""
        existing = self.db.execute(
            "SELECT id FROM ai_queue WHERE asset_id=? AND task_type=? AND status='pending'",
            (asset_id, task_type),
        )
        if existing:
            return
        self.db.execute(
            "INSERT INTO ai_queue (asset_id, task_type) VALUES (?,?)",
            (asset_id, task_type),
        )

    def process_queue(self) -> int:
        """Process all pending AI tasks. Retries on failure up to MAX_RETRIES
        within the same loop iteration. Returns count of completed tasks."""
        print("[AIWorker] process_queue() called", flush=True)
        try:
            rows = self.db.execute(
                "SELECT aq.id, aq.asset_id, aq.task_type, a.path, a.asset_type, aq.attempt "
                "FROM ai_queue aq JOIN assets a ON aq.asset_id = a.id "
                "WHERE aq.status = 'pending' ORDER BY aq.id"
            )
            print(f"[AIWorker] query returned {len(rows)} rows", flush=True)
        except Exception as e:
            print(f"[AIWorker] query ERROR: {e}", flush=True)
            import traceback
            traceback.print_exc()
            return 0
        if not rows:
            return 0
        print(f"[AIWorker] 发现 {len(rows)} 个待处理任务", flush=True)
        count = 0
        for row in rows:
            name = self.db.execute(
                "SELECT filename FROM assets WHERE id=?", (row["asset_id"],)
            )[0]["filename"]
            asset_type = row["asset_type"]
            task_type = row["task_type"]
            attempt = row["attempt"] or 0

            # Retry loop: keep trying the same task until success or max retries
            while True:
                self.db.execute(
                    "UPDATE ai_queue SET status='processing', attempt=? WHERE id=?",
                    (attempt, row["id"]),
                )
                print(f"[AIWorker] 分析中: {name} ({task_type}/{asset_type})"
                      f"{f' (第{attempt+1}次尝试)' if attempt > 0 else ''}", flush=True)
                try:
                    if task_type == "vision":
                        if asset_type == "video":
                            self._process_video(row["asset_id"], row["path"])
                        else:
                            self._process_vision(row["asset_id"], row["path"])
                    elif task_type == "text":
                        self._process_text(row["asset_id"], row["path"])
                    elif task_type == "transcribe":
                        self._process_transcribe(row["asset_id"], row["path"])
                    elif task_type == "embedding":
                        self._process_embedding(row["asset_id"])
                    self.db.execute(
                        "UPDATE ai_queue SET status='done', attempt=? WHERE id=?",
                        (attempt, row["id"]),
                    )
                    # Update analyzed_at timestamp
                    from datetime import datetime
                    self.db.execute(
                        "UPDATE assets SET analyzed_at=? WHERE id=?",
                        (datetime.now().isoformat(), row["asset_id"]),
                    )
                    print(f"[AIWorker] 完成: {name}", flush=True)
                    # Enqueue embedding if result is sufficient for vectorization
                    if task_type in ("vision", "text", "transcribe", "video_summary"):
                        self._enqueue_embedding(row["asset_id"])
                    break  # success
                except Exception as e:
                    attempt += 1
                    if attempt < self.MAX_RETRIES:
                        print(f"[AIWorker] 重试 {attempt}/{self.MAX_RETRIES}: {name} — {e}", flush=True)
                        import time; time.sleep(2)
                    else:
                        self.db.execute(
                            "UPDATE ai_queue SET status='failed', error=?, attempt=? WHERE id=?",
                            (str(e), attempt, row["id"]),
                        )
                        print(f"[AIWorker] 失败: {name} — {e}", flush=True)
                        break
            count += 1
        return count

    def _process_vision(self, asset_id: int, path: str) -> None:
        adapter = self._get_adapter("vision")
        analyzer = VisionAnalyzer(adapter=adapter, prompt_config=self._pc())
        result = analyzer.analyze(path)
        name = self.db.execute("SELECT filename FROM assets WHERE id=?", (asset_id,))[0]["filename"]
        print(f"[AIWorker] vision result for {name}: desc={bool(result.get('description'))}, tags={len(result.get('tags',[]))}, ocr={bool(result.get('ocr_text'))}", flush=True)
        if result.get("description"):
            self.db.execute(
                "UPDATE assets SET visual_description=? WHERE id=?",
                (result["description"], asset_id),
            )
        if result.get("ocr_text"):
            self.db.execute(
                "UPDATE assets SET ocr_text=? WHERE id=?",
                (result["ocr_text"], asset_id),
            )
        if result.get("tags"):
            self._link_tags(asset_id, result["tags"], source="auto")
        if result.get("search_terms"):
            self._save_search_terms(asset_id, result["search_terms"])

    def _process_video(self, asset_id: int, path: str) -> None:
        """Analyze a video using multi-frame sampling."""
        import tempfile
        num_frames = self.config.get("ai.video_frames") or 1
        frame_dir = tempfile.mkdtemp()
        # Debug: save frames to Desktop if QM_DEBUG_FRAMES is set
        _debug_dir = None
        if os.environ.get("QM_DEBUG_FRAMES"):
            import shutil, time
            _debug_dir = os.path.join(os.path.expanduser("~/Desktop/test_frame"), f"{os.path.basename(path)}_{time.strftime('%Y%m%d_%H%M%S')}")
            os.makedirs(_debug_dir, exist_ok=True)
        frames = extract_video_frames(path, frame_dir, num_frames)
        if _debug_dir and frames:
            for fp in frames:
                shutil.copy2(fp, os.path.join(_debug_dir, os.path.basename(fp)))
        if frames:
            adapter = self._get_adapter("video_vision")
            if len(frames) == 1:
                analyzer = VisionAnalyzer(adapter=adapter, prompt_config=self._pc(), prompt_type="video_vision")
                merged = analyzer.analyze(frames[0])
            else:
                analyzer = VisionAnalyzer(adapter=adapter, prompt_config=self._pc(), prompt_type="video_vision")
                merged = analyzer.analyze_multi(frames)
            if merged.get("description"):
                self.db.execute(
                    "UPDATE assets SET visual_description=? WHERE id=?",
                    (merged["description"], asset_id),
                )
            if merged.get("ocr_text"):
                self.db.execute(
                    "UPDATE assets SET ocr_text=? WHERE id=?",
                    (merged["ocr_text"], asset_id),
                )
            if merged.get("tags"):
                self._link_tags(asset_id, merged["tags"], source="auto")
            if merged.get("search_terms"):
                self._save_search_terms(asset_id, merged["search_terms"])
        # Try generating combined summary if transcript exists
        self._try_generate_video_summary(asset_id)

    def _process_transcribe(self, asset_id: int, path: str) -> None:
        """Transcribe audio/video file, then run speech analysis."""
        transcript = self._transcriber.transcribe(path)
        if transcript:
            self.db.execute(
                "UPDATE assets SET transcript=? WHERE id=?",
                (transcript, asset_id),
            )
            name = self.db.execute(
                "SELECT filename FROM assets WHERE id=?", (asset_id,)
            )[0]["filename"]
            print(f"[AIWorker] 转录完成: {name} ({len(transcript)}字)", flush=True)
            # Run speech analysis on transcript
            adapter = self._get_adapter("speech")
            analyzer = TextAnalyzer(adapter=adapter, prompt_config=self._pc())
            result = analyzer.analyze_speech(transcript)
            if result.get("summary"):
                self.db.execute(
                    "UPDATE assets SET ai_summary=? WHERE id=?",
                    (result["summary"], asset_id),
                )
            if result.get("tags"):
                self._link_tags(asset_id, result["tags"], source="auto")
            atype = self.db.execute("SELECT asset_type FROM assets WHERE id=?", (asset_id,))[0]["asset_type"]
            if result.get("search_terms") and atype == "audio":
                self._save_search_terms(asset_id, result["search_terms"])
            # Try generating combined summary if vision is also done
            self._try_generate_video_summary(asset_id)

    def _try_generate_video_summary(self, asset_id: int) -> None:
        """Generate combined summary if both speech and vision are done for a video."""
        rows = self.db.execute(
            "SELECT asset_type, visual_description, ai_summary FROM assets WHERE id=?",
            (asset_id,),
        )
        if not rows:
            return
        asset = rows[0]
        if asset["asset_type"] != "video":
            return
        if not asset["visual_description"]:
            return
        # Silent video: copy visual_description to video_summary
        has_transcript = bool(self.db.execute("SELECT transcript FROM assets WHERE id=?", (asset_id,))[0]["transcript"]) if self.db.execute("SELECT 1 FROM assets WHERE id=?", (asset_id,)) else False
        if not has_transcript:
            self.db.execute("UPDATE assets SET video_summary=? WHERE id=?", (asset["visual_description"], asset_id))
            print(f"[AIWorker] 无声视频: video_summary=visual_description", flush=True)
            return

        adapter = self._get_adapter("video_summary")
        # Generate combined summary via Ollama
        if self._pc():
            base = self._pc().get_prompt("video_summary")
        else:
            from quickmedia.prompt_config import DEFAULT_PROMPTS
            base = "".join(DEFAULT_PROMPTS["video_summary"]["default"])
        prompt = (
            f"{base}\n\n"
            f"画面内容：{asset['visual_description']}\n\n"
            f"语音内容：{asset['ai_summary']}\n\n"
            "综合总结："
        )
        response = adapter.chat(prompt)
        if response:
            import json as _json
            try:
                parsed = _json.loads(response) if isinstance(response, str) else response
                if isinstance(parsed, dict):
                    if parsed.get("video_summary"):
                        self.db.execute("UPDATE assets SET video_summary=? WHERE id=?", (parsed["video_summary"], asset_id))
                    if parsed.get("tags"):
                        self._link_tags(asset_id, parsed["tags"], source="auto")
                    if parsed.get("search_terms"):
                        self._save_search_terms(asset_id, parsed["search_terms"])
            except Exception:
                self.db.execute("UPDATE assets SET video_summary=? WHERE id=?", (response.strip(), asset_id))
            name = self.db.execute("SELECT filename FROM assets WHERE id=?", (asset_id,))[0]["filename"]
            print(f"[AIWorker] 综合总结完成: {name}", flush=True)

    def _process_text(self, asset_id: int, path: str) -> None:
        G = "[92m"; R = "[0m"
        import os as _os
        ext = _os.path.splitext(path)[1].lower()

        # Check if model supports native document upload for this format
        # Currently only OpenAI direct (not OpenRouter) and Gemini support file upload
        binding = self._registry.get_task_binding("text")
        if binding and binding["provider"] in ("openai",):
            from quickmedia.providers import ProviderRegistry
            model_info = self._registry.get_model_info(binding["provider"], binding["model"])
            doc_formats = (model_info.get("capabilities") or {}).get("document", []) if model_info else []
            supported_format = ext.lstrip(".").upper()
            if any(f.upper() == supported_format or f.lower() == ext.lstrip(".") for f in doc_formats):
                adapter = self._get_adapter("text")
                analyzer = TextAnalyzer(adapter=adapter, prompt_config=self._pc())
                print(f"{G}[文档分析] 传文件模式: {_os.path.basename(path)} → {binding['model']}{R}", flush=True)
                result = analyzer.analyze_file(path)
                if result.get("summary"):
                    self.db.execute("UPDATE assets SET ai_summary=? WHERE id=?", (result["summary"], asset_id))
                if result.get("tags"):
                    self._link_tags(asset_id, result["tags"], source="auto")
                if result.get("search_terms"):
                    self._save_search_terms(asset_id, result["search_terms"])
                return

        # Fall back to text extraction
        text = self._read_text(path)
        if text:
            print(f"{G}[文档分析] 提取文字模式: {_os.path.basename(path)} ({len(text)}字){R}", flush=True)
            adapter = self._get_adapter("text")
            analyzer = TextAnalyzer(adapter=adapter, prompt_config=self._pc())
            result = analyzer.analyze(text)
            if result.get("summary"):
                self.db.execute("UPDATE assets SET ai_summary=? WHERE id=?", (result["summary"], asset_id))
            if result.get("tags"):
                self._link_tags(asset_id, result["tags"], source="auto")
            if result.get("search_terms"):
                self._save_search_terms(asset_id, result["search_terms"])


    def _read_text(self, path: str) -> str:
        _, ext = os.path.splitext(path)
        ext = ext.lower()
        if ext in {".txt", ".md"}:
            with open(path, "r", errors="replace") as f:
                return f.read()
        if ext == ".csv":
            import csv, io
            with open(path, "r", errors="replace") as f:
                reader = csv.reader(f)
                rows = list(reader)
                if not rows:
                    return ""
                max_cols = max(len(r) for r in rows)
                # Format as aligned table
                lines = []
                for row in rows:
                    padded = row + [""] * (max_cols - len(row))
                    lines.append(" | ".join(padded))
                return "\n".join(lines)
        if ext == ".json":
            import json
            with open(path, "r", errors="replace") as f:
                data = json.load(f)
                return json.dumps(data, ensure_ascii=False, indent=2)
        if ext == ".xlsx":
            import io
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                parts = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        rows.append(" | ".join(str(c) if c is not None else "" for c in row))
                    if rows:
                        parts.append(f"=== Sheet: {name} ===\n" + "\n".join(rows[:500]))
                wb.close()
                return "\n\n".join(parts)
            except ImportError:
                return ""
        if ext == ".pdf":
            try:
                import fitz
                doc = fitz.open(path)
                text = ""
                for page in doc:
                    text += page.get_text()
                doc.close()
                return text
            except ImportError:
                pass
        return ""

    def _link_tags(self, asset_id: int, tag_names: list[str], source: str = "auto") -> None:
        for name in tag_names:
            rows = self.db.execute("SELECT id FROM tags WHERE name=?", (name,))
            if rows:
                tag_id = rows[0]["id"]
            else:
                cursor = self.db.conn.execute(
                    "INSERT INTO tags (name) VALUES (?)", (name,)
                )
                self.db.conn.commit()
                tag_id = cursor.lastrowid
            existing = self.db.execute(
                "SELECT 1 FROM asset_tags WHERE asset_id=? AND tag_id=?",
                (asset_id, tag_id),
            )
            if not existing:
                self.db.execute(
                    "INSERT INTO asset_tags (asset_id, tag_id, source) VALUES (?,?,?)",
                    (asset_id, tag_id, source),
                )

    def _save_search_terms(self, asset_id: int, search_terms: list[str]) -> None:
        """Store search_terms, replacing any existing."""
        if not search_terms:
            return
        self.db.execute("DELETE FROM asset_search_terms WHERE asset_id=?", (asset_id,))
        self.db.conn.commit()
        count = 0
        for term in search_terms:
            term = term.strip()
            if term:
                try:
                    self.db.execute(
                        "INSERT INTO asset_search_terms (asset_id, term) VALUES (?,?)",
                        (asset_id, term),
                    )
                    count += 1
                except Exception:
                    pass
        if count > 0:
            print(f"[search_terms] saved {count} terms for asset {asset_id}", flush=True)

    def _enqueue_embedding(self, asset_id: int) -> None:
        """Enqueue an embedding task if not already queued for this asset."""
        existing = self.db.execute(
            "SELECT 1 FROM ai_queue WHERE asset_id=? AND task_type='embedding'",
            (asset_id,),
        )
        if not existing:
            self.db.execute(
                "INSERT INTO ai_queue (asset_id, task_type) VALUES (?, 'embedding')",
                (asset_id,),
            )

    def _load_api_key(self, provider_name: str) -> str:
        """Load API key for a provider from env dict."""
        key = self._env.get(f"{provider_name.upper()}_API_KEY", "")
        if provider_name == "ollama":
            return key or "ollama"
        return key

    def _process_embedding(self, asset_id: int) -> None:
        """Generate and store embedding for a single asset."""
        from quickmedia.providers import ProviderRegistry

        # Get asset data
        rows = self.db.execute("SELECT * FROM assets WHERE id=?", (asset_id,))
        if not rows:
            raise ValueError(f"Asset {asset_id} not found")
        asset = dict(rows[0])
        tags = self.db.get_asset_tags(asset_id)
        asset["tags"] = [dict(t) for t in tags]

        # Get embedding adapter
        binding = self._registry.get_task_binding("embedding")
        if not binding:
            return  # no embedding model configured

        url = self._registry.get_provider_url(binding["provider"])
        if not url:
            return

        if binding["provider"] == "ollama":
            adapter = EmbeddingAdapter(base_url=url, model=binding["model"])
        else:
            api_key = self._load_api_key(binding["provider"])
            adapter = OpenAiEmbeddingAdapter(base_url=url, api_key=api_key, model=binding["model"])

        analyzer = EmbeddingAnalyzer(adapter=adapter)

        # Embed each search_term individually
        terms = [r["term"] for r in self.db.execute("SELECT term FROM asset_search_terms WHERE asset_id=?", (asset_id,))]
        # Also include tag names as vectors
        tag_names = [r["name"] for r in self.db.execute(
            "SELECT t.name FROM asset_tags at JOIN tags t ON at.tag_id=t.id WHERE at.asset_id=?", (asset_id,)
        )]
        terms = list(set(terms + tag_names))  # deduplicated
        if not terms:
            return
        chroma_path = os.path.join(self._config_dir, "chroma_db")
        store = ChromaStore(persist_path=chroma_path)
        # Clear old search vectors for this asset
        store.delete(asset_id)
        for idx, term in enumerate(terms):
            vector = analyzer.embed(term)
            store.add(asset_id, "search", vector, term_index=idx, term_text=term)
            print(f"[Embedding] {term} -> dim={len(vector)}", flush=True)
        name_rows = self.db.execute("SELECT filename FROM assets WHERE id=?", (asset_id,))
        print(f"[Embedding] 完成: {name_rows[0]['filename'] if name_rows else asset_id} ({len(terms)} terms)", flush=True)
